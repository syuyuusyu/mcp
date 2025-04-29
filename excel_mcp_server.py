from fastmcp import FastMCP
import openpyxl
import pandas as pd
import os

# 创建 MCP 服务器实例
mcp = FastMCP("Excel Server")

@mcp.tool()
def read_sheet_names(fileAbsolutePath: str) -> list[str]:
    """读取 Excel 文件中的所有工作表名称。
    
    Args:
        fileAbsolutePath: Excel 文件的绝对路径
    Returns:
        工作表名称列表
    """
    if not os.path.exists(fileAbsolutePath):
        raise ValueError("File not found")
        
    workbook = openpyxl.load_workbook(fileAbsolutePath, data_only=False)
    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names

@mcp.tool()
def read_sheet_data(fileAbsolutePath: str, sheetName: str, range: str = None) -> list[list]:
    """读取 Excel 工作表中的数据。
    
    Args:
        fileAbsolutePath: Excel 文件的绝对路径
        sheetName: 工作表名称
        range: 可选的单元格范围，例如 "A1:C10"
    Returns:
        二维数组形式的数据
    """
    if not os.path.exists(fileAbsolutePath):
        raise ValueError("File not found")
        
    df = pd.read_excel(fileAbsolutePath, sheet_name=sheetName)
    return df.values.tolist()

@mcp.tool()
def read_sheet_formula(fileAbsolutePath: str, sheetName: str, range: str = None) -> list[list[str]]:
    """读取 Excel 工作表中的公式。
    
    Args:
        fileAbsolutePath: Excel 文件的绝对路径
        sheetName: 工作表名称
        range: 可选的单元格范围，例如 "A1:C10"
    Returns:
        二维数组形式的公式
    """
    if not os.path.exists(fileAbsolutePath):
        raise ValueError("File not found")
        
    workbook = openpyxl.load_workbook(fileAbsolutePath, data_only=False)
    sheet = workbook[sheetName]
    
    formulas = []
    for row in sheet.iter_rows():
        row_formulas = []
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                row_formulas.append(cell.value)
            else:
                row_formulas.append(None)
        formulas.append(row_formulas)
    
    workbook.close()
    return formulas

@mcp.tool()
def write_sheet_data(fileAbsolutePath: str, sheetName: str, data: list[list], range: str = None) -> bool:
    """写入数据到 Excel 工作表。
    
    Args:
        fileAbsolutePath: Excel 文件的绝对路径
        sheetName: 工作表名称
        data: 要写入的二维数组数据
        range: 可选的目标单元格范围，例如 "A1:C10"
    Returns:
        是否写入成功
    """
    df = pd.DataFrame(data)
    with pd.ExcelWriter(fileAbsolutePath, engine='openpyxl', mode='a' if os.path.exists(fileAbsolutePath) else 'w') as writer:
        df.to_excel(writer, sheet_name=sheetName, index=False, header=False)
    return True

@mcp.tool()
def write_sheet_formula(fileAbsolutePath: str, sheetName: str, formulas: list[list[str]], range: str = None) -> bool:
    """写入公式到 Excel 工作表。
    
    Args:
        fileAbsolutePath: Excel 文件的绝对路径
        sheetName: 工作表名称
        formulas: 要写入的二维数组公式
        range: 可选的目标单元格范围，例如 "A1:C10"
    Returns:
        是否写入成功
    """
    workbook = openpyxl.load_workbook(fileAbsolutePath) if os.path.exists(fileAbsolutePath) else openpyxl.Workbook()
    
    if sheetName not in workbook.sheetnames:
        workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    
    for i, row in enumerate(formulas):
        for j, formula in enumerate(row):
            if formula:
                sheet.cell(row=i+1, column=j+1, value=formula)
    
    workbook.save(fileAbsolutePath)
    workbook.close()
    return True

@mcp.tool()
def create_excel_file(fileAbsolutePath: str, title: list[str], data: list[list[any]]) -> bool:
    """创建新的 Excel 文件并写入数据。
    
    Args:
        fileAbsolutePath: 要创建的 Excel 文件的绝对路径
        title: 表头列表，例如 ["姓名", "年龄", "成绩"]
        data: 要写入的数据，二维数组，每个内部数组代表一行数据
    Returns:
        是否创建成功
    """
    try:
        # 如果文件已存在，先删除
        if os.path.exists(fileAbsolutePath):
            os.remove(fileAbsolutePath)
            
        # 创建 DataFrame
        df = pd.DataFrame(data, columns=title)
        
        # 写入 Excel 文件
        df.to_excel(fileAbsolutePath, index=False, sheet_name='Sheet1')
        return True
    except Exception as e:
        raise ValueError(f"Failed to create Excel file: {str(e)}")

if __name__ == "__main__":
    # 使用默认的 stdio 传输运行服务器
    mcp.run() 